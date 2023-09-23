"""Imports xls statements from pumb, optained via request to online support.

The header is as follows:
Дата операції	Тип операції	Сума у валюті операції	Опис	Дата списання	Сума операції	Комісія
"""

import datetime
import os

import beangulp
import dateutil.parser
import openpyxl
from beancount.core import data, flags
from beancount.core.number import D

from uabean.importers.mixins import IdentifyMixin


class Importer(IdentifyMixin, beangulp.Importer):
    FLAG = flags.FLAG_OKAY
    matchers = [
        ("filename", "pumb"),
        ("mime", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
    ]
    DATE_COL = 0
    TYPE_COL = 1
    TRANSACTION_CURRENCY_AMOUNT_COL = 2
    DESCRIPTION_COL = 3
    SETTLE_DATE_COL = 4
    ACCOUNT_CURRENCY_AMOUNT_COL = 5
    FEE_COL = 6
    POS_PAYMENT_STR = "Oперация покупки через POS-терминал"
    CASHBACK_STR = "Виплата винагороди кешбек ПУМБ"

    def __init__(
        self,
        account_config,
        *args,
        fee_account="Expenses:Fees:Pumb",
        cashback_income_account="Income:Cashback:Pumb",
        **kwargs,
    ):
        self.account_config = account_config
        self.fee_account = fee_account
        self.cashback_income_account = cashback_income_account
        super().__init__(*args, **kwargs)

    def datetime_from_row(self, row):
        return dateutil.parser.parse(row[self.DATE_COL].value, dayfirst=True)

    @classmethod
    def get_amount(cls, cell, currency=None):
        if currency is None:
            n, currency = cell.value.rsplit(" ", 1)
        else:
            n = cell.value
        return data.Amount(D(n.replace(" ", "").replace(",", ".")), currency)

    def account(self, _):
        return "pumb"

    def extract(self, filename, existing_entries=None):
        entries = []
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.worksheets[0]
        assert "Виписка по операціям" in sheet.cell(7, 1).value
        account_currency = os.path.basename(filename.split("-")[1]).upper()
        account = self.account_config[account_currency]
        date_row_seen = False
        balance_end_date = None
        balance_entry = None
        for i, row in enumerate(sheet.iter_rows()):
            if not date_row_seen:
                if row[0].value == "Дата операції":
                    date_row_seen = True
                if row[0].value == "Період":
                    balance_end_date = dateutil.parser.parse(
                        row[1].value.split(" - ")[-1], dayfirst=True
                    )
                if row[0].value == "Баланс на кінець періоду":
                    balance_entry = data.Balance(
                        data.new_metadata(filename, i + 1),
                        balance_end_date.date() + datetime.timedelta(days=1),
                        account,
                        self.get_amount(row[1], account_currency),
                        None,
                        None,
                    )
                continue
            if row[0].value == "ВСЬОГО":
                break
            meta = data.new_metadata(filename, i + 1)
            entries.append(self.entry_from_row(meta, account, row))
        if balance_entry is not None:
            entries.append(balance_entry)
        return entries

    def entry_from_row(self, meta, account, row):
        dt = self.datetime_from_row(row)
        meta["time"] = dt.strftime("%H:%M")
        narration = row[self.DESCRIPTION_COL].value
        if self.POS_PAYMENT_STR in narration:
            narration = narration.split("; ")[-1]
        account_currency_amount = self.get_amount(row[self.ACCOUNT_CURRENCY_AMOUNT_COL])
        transaction_currency_amount = self.get_amount(
            row[self.TRANSACTION_CURRENCY_AMOUNT_COL]
        )
        postings = [
            data.Posting(account, account_currency_amount, None, None, None, None)
        ]
        if (
            account_currency_amount.currency != transaction_currency_amount.currency
            and account_currency_amount.number != transaction_currency_amount.number
        ):
            meta["converted"] = f"{-transaction_currency_amount}"
        elif account_currency_amount.number != transaction_currency_amount.number:
            fee_amount = data.Amount(
                abs(
                    account_currency_amount.number - transaction_currency_amount.number
                ),
                account_currency_amount.currency,
            )
            postings.append(
                data.Posting(self.fee_account, fee_amount, None, None, None, None)
            )
        if self.CASHBACK_STR in narration:
            postings.append(
                data.Posting(
                    self.cashback_income_account,
                    -account_currency_amount,
                    None,
                    None,
                    None,
                    None,
                )
            )
            narration = self.CASHBACK_STR
        return data.Transaction(
            meta,
            dt.date(),
            self.FLAG,
            None,
            narration,
            data.EMPTY_SET,
            data.EMPTY_SET,
            postings,
        )


def get_test_importer():
    return Importer(
        account_config={
            "UAH": "Assets:Pumb:Cash:UAH",
            "EUR": "Assets:EUR:Pumb",
        }
    )


if __name__ == "__main__":
    from beangulp.testing import main

    main(get_test_importer())
