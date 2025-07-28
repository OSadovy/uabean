"""Imports xlsx statements from privatbank, optained via privat24 web interface.

The header is as follows:
Дата;Категорія;Картка;Опис операції;Сума в валюті картки;Валюта картки;Сума в валюті транзакції;Валюта транзакції;Залишок на кінець періоду;Валюта залишку

Note: Date column contains combined date and time in format "DD.MM.YYYY HH:MM:SS"
"""

import datetime

import beangulp
import dateutil.parser
import openpyxl
from beancount.core import data, flags
from beancount.core.number import D

from uabean.importers.mixins import IdentifyMixin


class Importer(IdentifyMixin, beangulp.Importer):
    FLAG = flags.FLAG_OKAY
    matchers = [
        ("content", "Виписка з Ваших карток"),
        ("mime", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
    ]
    
    @staticmethod
    def converter(filename):
        """Convert XLSX to text for content matching."""
        try:
            workbook = openpyxl.load_workbook(filename, read_only=True, data_only=True)
            try:
                sheet = workbook.worksheets[0]
                # Get first cell content which contains the header
                first_cell = sheet.cell(row=1, column=1).value
                if first_cell:
                    return str(first_cell)
            finally:
                workbook.close()
        except Exception as e:
            # Log the error for debugging in test environments
            import sys
            print(f"Error in privatbank_xlsx converter: {e}", file=sys.stderr)
        return ""
    unknown_account = "Assets:Unknown"
    DATE_COL = 0
    CATEGORY_COL = 1
    CARD_COL = 2
    DESCRIPTION_COL = 3
    CARD_CURRENCY_AMOUNT_COL = 4
    CARD_CURRENCY_COL = 5
    TRANSACTION_AMOUNT_COL = 6
    TRANSACTION_CURRENCY_COL = 7
    BALANCE_COL = 8
    BALANCE_CURRENCY_COL = 9
    CURRENCY_MAP = {"грн": "UAH", "дол": "USD", "євро": "EUR", "PLN": "PLN"}

    def __init__(
        self,
        card_to_account_map,
        *args,
        fee_account="Expenses:Fees:Privatbank",
        **kwargs,
    ):
        self.card_to_account_map = card_to_account_map
        self.fee_account = fee_account
        super().__init__(*args, **kwargs)

    def date_from_row(self, row):
        # Parse combined date/time format like "23.07.2025 00:37:19"
        datetime_str = row[self.DATE_COL].value
        return dateutil.parser.parse(datetime_str, dayfirst=True)

    @classmethod
    def get_currency(cls, cell):
        return cls.CURRENCY_MAP.get(cell.value, cell.value)

    @staticmethod
    def get_number(cell):
        return D(str(cell.value))

    def account(self, _):
        return "privatbank_xlsx"

    def extract(self, filename, existing_entries=None):
        entries = []
        workbook = openpyxl.load_workbook(filename, read_only=True, data_only=True)
        try:
            sheet = workbook.worksheets[0]
            assert "Виписка з Ваших карток за період" in sheet.cell(row=1, column=1).value
            max_date = None
            max_row = None
            # Convert sheet rows to list for easier access (openpyxl uses 1-based indexing)
            rows = list(sheet.iter_rows())
            for nrow in range(2, len(rows)):  # Start from row 3 (index 2) like in xlrd version
                row = rows[nrow]
                meta = data.new_metadata(filename, nrow + 1)  # +1 for 1-based line numbers
                entries.append(self.entry_from_row(meta, row))
                date = self.date_from_row(row)
                if max_date is None or date > max_date:
                    max_date = date
                    max_row = row
            if max_row is not None:
                amount = data.Amount(
                    self.get_number(max_row[self.BALANCE_COL]),
                    self.get_currency(max_row[self.BALANCE_CURRENCY_COL]),
                )
                entries.append(
                    data.Balance(
                        data.new_metadata(filename, 0),
                        max_date.date() + datetime.timedelta(days=1),
                        self.card_to_account_map[max_row[self.CARD_COL].value],
                        amount,
                        None,
                        None,
                    )
                )
        finally:
            workbook.close()
        return entries

    def entry_from_row(self, meta, row):
        dt = self.date_from_row(row)
        # Extract time from the datetime for metadata
        meta["time"] = dt.strftime("%H:%M:%S")
        meta["category"] = row[self.CATEGORY_COL].value
        account = self.card_to_account_map.get(
            row[self.CARD_COL].value, self.unknown_account
        )
        num = self.get_number(row[self.TRANSACTION_AMOUNT_COL])
        currency = self.get_currency(row[self.TRANSACTION_CURRENCY_COL])
        card_num = self.get_number(row[self.CARD_CURRENCY_AMOUNT_COL])
        card_currency = self.get_currency(row[self.CARD_CURRENCY_COL])
        postings = [
            data.Posting(
                account, data.Amount(card_num, card_currency), None, None, None, None
            )
        ]
        if currency != card_currency:
            meta["converted"] = f"{num} {currency}"
        elif abs(card_num) != num:
            fee_amount = data.Amount(abs(card_num) - num, currency)
            postings.append(
                data.Posting(self.fee_account, fee_amount, None, None, None, None)
            )
        return data.Transaction(
            meta,
            dt.date(),
            self.FLAG,
            None,
            row[self.DESCRIPTION_COL].value,
            data.EMPTY_SET,
            data.EMPTY_SET,
            postings,
        )


def get_test_importer():
    return Importer(
        {
            "1234": "Assets:Privatbank:Universal",
            "5678": "Assets:Privatbank:Social",
            "8345 **** **** 2284": "Assets:Privatbank:Test",
        }
    )


if __name__ == "__main__":
    from beangulp.testing import main

    main(get_test_importer())